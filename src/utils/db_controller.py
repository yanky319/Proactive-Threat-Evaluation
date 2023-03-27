import os
import logging

import openpyxl as excel_db

from src.config import SHEET_NAME, DBFields, LOGGER_NAME

logger = logging.getLogger(LOGGER_NAME)


class DB:

    def __init__(self, path):
        self.db_path = path
        if os.path.isfile(path):
            logger.info(f'loading DB from {path}')
            self.wb = excel_db.load_workbook(path)
            self.ws = self.wb[SHEET_NAME]
        else:
            logger.info(f'creating DB at {path}')
            self.wb = excel_db.Workbook()  # if Excel file not exist create new
            self.wb['Sheet'].title = SHEET_NAME
            self.ws = self.wb[SHEET_NAME]
            self.update_row(1, *tuple(h.value for h in DBFields))

    def save(self):
        self.wb.save(self.db_path)

    def add_value(self, row, column, value):
        self.ws.cell(row=row, column=column).value = value
        self.save()

    def add_row(self, *values):
        row = self.ws.max_row + 1
        self.update_row(row, *values)

    def update_row(self, row, *values):
        for column in range(len(values)):
            self.ws.cell(row=row, column=column + 1).value = values[column]
        self.save()

    def get_row_by_number(self, row_number):
        """Returns a list containing the values in the specified row of an Excel sheet."""

        return [cell.value for cell in self.ws[row_number]]

    def get_column(self, column_name):
        """Returns a list containing the values in the specified column of an Excel sheet."""

        column_number = self.get_row_by_number(1).index(column_name)
        column = []
        for row in self.ws.iter_rows(min_row=2, min_col=column_number, max_col=column_number):
            for cell in row:
                column.append(cell.value)
        return column

    def get_row_with_value(self, value):
        """Returns a list containing the values in the first row of an Excel sheet
        that includes the specified value."""
        values = []
        for row in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == value:
                    values = [cell.offset(column=col_index - cell.column).value for col_index in
                              range(1, self.ws.max_column + 1)]

        keys = self.get_row_by_number(1)
        return dict(zip(keys, values))


