import datetime
import os
import json
import requests
import logging
import traceback

from src.config import DEFAULT_LAST_DATE, LOGGER_NAME
from src.utils import validate_ip_address
import openpyxl as excel_db

logger = logging.getLogger(LOGGER_NAME)
root_path = r'C:\Users\uziel.shemesh\PycharmProjects\Proactive-Threat-Evaluation'  # TODO change the path to constant location of the excel db
db_path = os.path.join(root_path, 'IOCs_DB.xlsx')
Hashes_sheet_name = 'Hashes'


def get_workbook():
    try:
        wb = excel_db.load_workbook(db_path)
    except:
        wb = excel_db.Workbook()  # if Excel file not exist create new
    if len(wb.sheetnames) == 0:
        wb.create_sheet(Hashes_sheet_name)
    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == 'Sheet':
        wb['Sheet'].title = Hashes_sheet_name
    ws = wb[Hashes_sheet_name]
    ws.cell(row=1, column=1).value = 'SHA256'
    ws.cell(row=1, column=2).value = 'MD5'
    ws.cell(row=1, column=3).value = 'Date uploaded to vt'
    ws.cell(row=1, column=4).value = 'file type'
    ws.cell(row=1, column=5).value = 'file name'
    wb.save(db_path)
    return wb


def get_sheet(wb, blog_name):
    if blog_name in wb.sheetnames:
        # Get the sheets by name
        return wb[blog_name]
    else:
        wb.create_sheet(blog_name)
        ws = wb[blog_name]
        ws.cell(row=1, column=1).value = 'Scan date'
        ws.cell(row=1, column=2).value = 'Blog upload date'
        ws.cell(row=1, column=3).value = 'URL'
        ws.cell(row=1, column=4).value = 'SHA256'
        ws.cell(row=1, column=5).value = 'MD5'
        ws.cell(row=1, column=6).value = 'Existence_in_VT'
        ws.cell(row=1, column=7).value = 'Date_uploaded_to_vt'
        ws.cell(row=1, column=8).value = 'file_type'
        ws.cell(row=1, column=9).value = 'malware_bazaar'
        return ws


class Scraper:
    def __init__(self, base, start, last_blog_date, extractor, pdf_generator, upload, folder):
        self.base_url = base
        self.start_url = start
        self.blogs = dict()
        self.last_blog_date = last_blog_date if last_blog_date else DEFAULT_LAST_DATE
        self.extractor = extractor
        self.pdf_generator = pdf_generator
        self.folder = os.path.join(folder, self.__class__.__name__)
        self.upload = upload
        self.accept_cookies_text = None
        os.makedirs(self.folder, exist_ok=True)

        #  extract and store temporarily all hashes from excel for rapid search
        wb = get_workbook()
        ws = wb[Hashes_sheet_name]
        self.all_sha256_hashes = [row[0].value for row in ws.iter_rows(min_row=ws.min_row + 1, max_row=ws.max_row, max_col=1)]
        self.all_md5_hashes = [row[0].value for row in ws.iter_rows(min_row=ws.min_row + 1, max_row=ws.max_row, min_col=2, max_col=2)]

        with open(root_path + '\\vt_apikey') as vt_key:
            self.vt_api_key = vt_key.read().strip()

    def find_new_blogs(self):
        raise NotImplementedError

    def update_db(self, result, post, vt):
        wb = get_workbook()
        blog_name = self.__class__.__name__
        blog_sheet = get_sheet(wb, blog_name)
        hashes_sheet = get_sheet(wb, Hashes_sheet_name)
        for url_row in range(1, blog_sheet.max_row):
            if blog_sheet.cell(row=url_row, column=3).value == post[0]:
                sha256s = blog_sheet.cell(row=url_row, column=4).value
            else:
                continue
            if not result.get('sha256_hash'):
                continue
            for sha in result.get('sha256_hash'):
                if sha not in sha256s.split(';'):
                    blog_sheet.cell(row=url_row, column=4).value += f';{sha}'

        # save the data
        if len(vt) > 0:
            for dic in vt:
                hashes_row = hashes_sheet.max_row + 1
                hashes_sheet.cell(row=hashes_row, column=1).value = dic['sha256']
                hashes_sheet.cell(row=hashes_row, column=2).value = dic['md5']
                hashes_sheet.cell(row=hashes_row, column=3).value = dic['first_submission_date']
                hashes_sheet.cell(row=hashes_row, column=4).value = dic['file_type']
                hashes_sheet.cell(row=hashes_row, column=5).value = dic['meaningful_name']
        current_row = blog_sheet.max_row + 1
        blog_sheet.cell(row=current_row, column=1).value = datetime.date.today().strftime('%d/%m/%Y')
        try:
            blog_sheet.cell(row=current_row, column=2).value = post[1].strftime('%d/%m/%Y')
        except Exception as e:
            print(f'blog_name - {e}')
        blog_sheet.cell(row=current_row, column=3).value = post[0]
        try:
            blog_sheet.cell(row=current_row, column=4).value = ';'.join(result.get('sha256_hash'))
        except Exception as e:
            print(f'blog_name - No SHA256 in that post add from Virus total')  # TODO add from VT result
            blog_sheet.cell(row=current_row, column=4).value = ';'.join([dic['sha256'] for dic in vt])
        try:
            blog_sheet.cell(row=current_row, column=5).value = ';'.join(result.get('md5_hash'))
        except Exception as e:
            print(f'blog_name - No MD5 in that post add from Virus total')
            blog_sheet.cell(row=current_row, column=4).value = ';'.join([dic['md5'] for dic in vt])

        # TODO insert hash data from VT to blog_sheet, and if it's not in VT insert 'False' to column 6
        blog_sheet.cell(row=current_row, column=6).value = 'True'  # TODO check existing in VT
        blog_sheet.cell(row=current_row, column=7).value = 'date_uploaded_to_vt'  # TODO check upload date to VT
        blog_sheet.cell(row=current_row, column=8).value = 'file_type'  # TODO check file type from VT
        blog_sheet.cell(row=current_row, column=9).value = 'malware_bazaar'  # TODO get data from 'malware bazaar'
        wb.save(db_path)

    def check_virus_total(self, result):
        base_vt_url = "https://www.virustotal.com/api/v3/files/{end_url}"
        headers = {
            "accept": "application/json",
            "x-apikey": self.vt_api_key
        }
        res = []
        if not result.get('sha256_hash') and not result.get('md5_hash'):
            return res
        elif not result.get('sha256_hash'):
            hashes = result.get('md5_hash')
        elif not result.get('md5_hash'):
            hashes = result.get('sha256_hash')
        else:
            if len(result.get('md5_hash')) > result.get('sha256_hash'):
                hashes = result.get('md5_hash')
            else:
                hashes = result.get('sha256_hash')
        hashes_to_check = []
        for current_hash in hashes:
            if current_hash in self.all_sha256_hashes:
                continue
            else:
                hashes_to_check.append(current_hash)
        for current_hash in hashes_to_check:
            response = requests.get(base_vt_url.format(end_url=current_hash), headers=headers)
            if response.status_code != 200:
                continue
            y = json.loads(response.text)['data']
            file_type = y['type']
            first_submission_date_unix_timestamp = y['attributes']['first_submission_date']
            first_submission_date = datetime.datetime.fromtimestamp(first_submission_date_unix_timestamp)
            md5 = y['attributes']['md5']
            sha256 = y['attributes']['sha256']
            meaningful_name = y['attributes']['meaningful_name']
            res.append({'sha256': sha256, 'md5': md5, 'first_submission_date': first_submission_date,
                    'file_type': file_type, 'meaningful_name': meaningful_name})
        return res

    def dump(self, name, content, pdf_bytes=None):
        logger.info(f'starting dumping for {name}')
        ioc_file_path = os.path.join(self.folder, f'{name}.json')
        with open(ioc_file_path, "w") as json_outfile:
            json.dump(content, json_outfile, indent=0)  # , default=set_default)

        if pdf_bytes:
            try:
                pdf_file_path = os.path.join(self.folder, f'{name}.pdf')
                with open(pdf_file_path, "wb") as pdf_outfile:
                    pdf_outfile.write(pdf_bytes)
            except Exception as e:
                tb = traceback.format_exc()
                logger.error(f'Cannot create PDF file for {name}\nexception: {e}\ntraceback: {tb}')

        if self.upload:
            pass
        # TODO upload json and PDF files

    @staticmethod
    def get_post_name(url):
        """
        get name of post
        :param url: link to the post
        :return: name of the post
        """
        raise NotImplementedError

    def scrape(self):
        self.find_new_blogs()

        for post in self.blogs.items():
            logger.info(f'starting scraping for {post[0]}')
            blog_name = self.get_post_name(post[0])
            pdf_bytes = None
            if self.pdf_generator:
                data, pdf_bytes = self.pdf_generator.generate_pdf_from_url(post[0], self.accept_cookies_text)
            else:
                page = requests.get(post[0])
                data = page.content.decode('utf-8')

            result = self.extractor.extract(src=data, defanged=False)
            vt = self.check_virus_total(result)  # TODO - insert VT result to the blog object
            self.update_db(result, post, vt)
            # result['ipv4'] = list(filter(validate_ip_address, result['ipv4']))  # validate ipv4 address
            # result['ipv6'] = list(filter(validate_ip_address, result['ipv6']))  # validate ipv6 address
            # result.pop('linux_path') if result.get('linux_path') else None
            # result.pop('windows_path') if result.get('windows_path') else None

            self.dump(name=blog_name, content=self.validate_and_cleanup(result), pdf_bytes=pdf_bytes)

    @staticmethod
    def validate_and_cleanup(result):
        """
        validate the ioc's that need validation, erase the ones that are not needed, and remove empty values
        :param result: result of IoCExtract.extract.
        :return: clean dictionary of validated ioc's
        """
        keys_to_pop = ['linux_path', 'windows_path']
        keys_to_validate = {'ipv4': validate_ip_address}

        for key in keys_to_validate.keys():
            if result.get(key):
                result[key] = list(filter(keys_to_validate[key], result[key]))

        for key in keys_to_pop:
            result.pop(key) if result.get(key) else None

        return {key: list(val) for key, val in result.items() if val}
